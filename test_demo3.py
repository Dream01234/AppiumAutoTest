from openpyxl import load_workbook
from appium import webdriver
import time
from appium.webdriver.common.touch_action import TouchAction
import Enlement
import method
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By

# caps = {}
# caps["platformName"] = "Android"
# caps["platformVersion"] = "7.1.2"
# caps["deviceName"] = "MI_9"
# caps["appPackage"] = "cn.smartinspection.combine"
# caps["appActivity"] = "cn.smartinspection.login.ui.activity.WelcomeGuideActivity"
# caps["ensureWebviewsHavePages"] = True
# caps["automationName"] = "Uiautomator2"
# driver = webdriver.Remote("http://localhost:4723/wd/hub",caps)


wb = load_workbook('D:/自动化模板.xlsx')
print(wb.sheetnames)
ws = wb['实测实量']
one_cel = ws.cell(row=1, column=2).value
two_cel = ws['A1'].value
three_cel = ws['B1'].value
four_cel = two_cel + three_cel

print(one_cel)
print(two_cel)
print(four_cel)
# Enlement.operation.find_uiautomator_tager_click(driver, one_cel, 3, 3)
